"""Builds an Excel or PDF export of an event's fund data (summary, expenses, vendors,
sponsorships) — used both for the authenticated download button and the public
share-link endpoint."""
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet


def _money(v) -> str:
    return f"Rs. {float(v):,.2f}"


async def fetch_fund_export_data(conn, event_id: str) -> dict:
    summary = await conn.fetchrow(
        "SELECT title, status, ticket_revenue, sponsorship_income, total_expenses, "
        "vendor_pool, net_balance, sponsor_count, complimentary_tickets "
        "FROM v_event_finance WHERE event_id = $1::uuid",
        event_id,
    )
    expenses = await conn.fetch(
        "SELECT ex.description, ex.category, ex.amount, ex.currency_code, "
        "u.name AS created_by_name, ex.created_at "
        "FROM event_expense ex JOIN users u ON u.id = ex.created_by "
        "WHERE ex.event_id = $1::uuid ORDER BY ex.created_at",
        event_id,
    )
    vendors = await conn.fetch(
        "SELECT v.name AS vendor_name, v.category, ev.stall_number, ev.fee_type, "
        "ev.fixed_fee, ev.revenue_share_pct, ev.actual_revenue, ev.status "
        "FROM event_vendor ev JOIN vendor v ON v.id = ev.vendor_id "
        "WHERE ev.event_id = $1::uuid ORDER BY v.name",
        event_id,
    )
    sponsorships = await conn.fetch(
        "SELECT s.organization_name, es.amount, es.currency_code, es.status, es.sponsored_at "
        "FROM event_sponsorship es JOIN sponsor s ON s.id = es.sponsor_id "
        "WHERE es.event_id = $1::uuid ORDER BY es.sponsored_at",
        event_id,
    )
    return {
        "summary": dict(summary) if summary else None,
        "expenses": [dict(r) for r in expenses],
        "vendors": [dict(r) for r in vendors],
        "sponsorships": [dict(r) for r in sponsorships],
    }


def build_xlsx(data: dict) -> bytes:
    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Summary"
    summary = data["summary"] or {}
    ws.append(["Event", summary.get("title", "")])
    ws.append(["Status", summary.get("status", "")])
    ws.append(["Ticket Revenue", float(summary.get("ticket_revenue", 0) or 0)])
    ws.append(["Sponsorship Income", float(summary.get("sponsorship_income", 0) or 0)])
    ws.append(["Total Expenses", float(summary.get("total_expenses", 0) or 0)])
    ws.append(["Vendor Pool", float(summary.get("vendor_pool", 0) or 0)])
    ws.append(["Net Balance", float(summary.get("net_balance", 0) or 0)])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        row[0].font = bold

    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Description", "Category", "Amount", "Currency", "Logged By", "Date"])
    for c in ws2[1]:
        c.font = bold
    for e in data["expenses"]:
        ws2.append([e["description"], e["category"], float(e["amount"]), e["currency_code"],
                    e["created_by_name"], e["created_at"].strftime("%Y-%m-%d")])

    ws3 = wb.create_sheet("Vendors")
    ws3.append(["Vendor", "Category", "Stall", "Fee Type", "Fixed Fee", "Revenue Share %", "Actual Revenue", "Status"])
    for c in ws3[1]:
        c.font = bold
    for v in data["vendors"]:
        ws3.append([v["vendor_name"], v["category"], v["stall_number"] or "", v["fee_type"],
                    float(v["fixed_fee"]), float(v["revenue_share_pct"]),
                    float(v["actual_revenue"]) if v["actual_revenue"] is not None else "",
                    v["status"]])

    ws4 = wb.create_sheet("Sponsorships")
    ws4.append(["Sponsor", "Amount", "Currency", "Status", "Date"])
    for c in ws4[1]:
        c.font = bold
    for s in data["sponsorships"]:
        ws4.append([s["organization_name"], float(s["amount"]), s["currency_code"],
                    s["status"], s["sponsored_at"].strftime("%Y-%m-%d")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = []

    summary = data["summary"] or {}
    story.append(Paragraph(f"Fund Report — {summary.get('title', '')}", styles["Title"]))
    story.append(Paragraph(f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    story.append(Spacer(1, 0.5 * cm))

    summary_rows = [
        ["Ticket Revenue", _money(summary.get("ticket_revenue", 0) or 0)],
        ["Sponsorship Income", _money(summary.get("sponsorship_income", 0) or 0)],
        ["Total Expenses", _money(summary.get("total_expenses", 0) or 0)],
        ["Vendor Pool", _money(summary.get("vendor_pool", 0) or 0)],
        ["Net Balance", _money(summary.get("net_balance", 0) or 0)],
    ]
    t = Table(summary_rows, colWidths=[6 * cm, 6 * cm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.8 * cm))

    if data["expenses"]:
        story.append(Paragraph("Expenses", styles["Heading2"]))
        rows = [["Description", "Category", "Amount", "Logged By"]]
        for e in data["expenses"]:
            rows.append([e["description"], e["category"], _money(e["amount"]), e["created_by_name"]])
        t = Table(rows, colWidths=[6 * cm, 3 * cm, 3 * cm, 4 * cm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.6 * cm))

    if data["sponsorships"]:
        story.append(Paragraph("Sponsorships", styles["Heading2"]))
        rows = [["Sponsor", "Amount", "Status"]]
        for s in data["sponsorships"]:
            rows.append([s["organization_name"], _money(s["amount"]), s["status"]])
        t = Table(rows, colWidths=[7 * cm, 4 * cm, 4 * cm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ]))
        story.append(t)

    doc.build(story)
    return buf.getvalue()
